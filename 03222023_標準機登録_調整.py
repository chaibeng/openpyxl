import tkinter as tk
from tkinter import filedialog
import openpyxl
import time #time

start_time = time.time() #time

# Create a tkinter window to prompt user to select files
root = tk.Tk()
root.withdraw()

# Ask user to select the first Excel file
file_path_1 = filedialog.askopenfilename(title="Select Intermediate_files.xlsx")

# Load the first workbook
wb1 = openpyxl.load_workbook(file_path_1)# Print out the existing worksheets

# Print the list of worksheet names
wb1_names = wb1.sheetnames
print("Existing worksheets:")
for i, name in enumerate(wb1_names):
    print(f"{i+1}. {name}")

# Prompt the user to select a worksheet
ws1_num = int(input("Enter the number of the worksheet (SAP data): "))
# Get the selected worksheet
ws1 = wb1[wb1_names[ws1_num-1]]

# Prompt the user to select a worksheet
ws4_num = int(input("Enter the number of the worksheet (RSJ Cal Standards): "))
# Get the selected worksheet
ws4 = wb1[wb1_names[ws4_num-1]]

# Prompt the user to select a worksheet
ws5_num = int(input("Enter the number of the worksheet (RSJ Cal Standards 2021-7-1): "))
# Get the selected worksheet
ws5 = wb1[wb1_names[ws5_num-1]]

print("-"*47)

'''-------------------------------------------------------------------------------------------------------------------'''
#調整

for i, row in enumerate(ws1.iter_rows(min_row=1,max_col=3,values_only=True),start=1):

    for j, row_WS in enumerate(ws5.iter_rows(min_row=2,min_col=1, max_col=9, values_only=True),start=2):

        if  row_WS[8] == row[2]:

            ID = ws5.cell(row=j,column=1).value #A
            print(ID)
            ws4.cell(row=i+1, column=1).value =ID #A
            print("-"*45)

            Order_Type= ws5.cell(row=j,column=13).value #M 
            print(Order_Type)
            ws4.cell(row=i+1, column=11).value = Order_Type #K
            print("-"*45)
            
            System= ws5.cell(row=j,column=11).value  #K
            print( System)
            ws4.cell(row=i+1, column=12).value =  System #L
            print("-"*45)

            Data_Update= ws5.cell(row=j,column=17).value #Q
            print( Data_Update)
            ws4.cell(row=i+1, column=13).value =  Data_Update #M
            print("-"*45)

            Data_Update= ws5.cell(row=j,column=18).value #R
            print( Data_Update)
            ws4.cell(row=i+1, column=14).value =  Data_Update #N
            print("-"*45)

            Data_Update= ws5.cell(row=j,column=26).value #Z
            print( Data_Update)
            ws4.cell(row=i+1, column=19).value =  Data_Update #S 
            print("-"*45)

'''-------------------------------------------------------------------------------------------------------------------'''
print("-"*47)

# Do some processing here... #time
end_time = time.time()
# Calculate the time elapsed
elapsed_time = end_time - start_time
# Convert elapsed time to minutes and seconds
minutes = int(elapsed_time // 60)
seconds = int(elapsed_time % 60)
print(f'The code took {minutes} minutes and {seconds} seconds to execute.')

def save_workbook(workbook, filename):
    workbook.save(filename)
save_workbook(wb1, "標準機登録_FINAL_DATA.xlsx")