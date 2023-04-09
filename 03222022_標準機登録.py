import tkinter as tk
from tkinter import filedialog
import openpyxl
import time #time

start_time = time.time() #time

# Create a tkinter window to prompt user to select files
root = tk.Tk()
root.withdraw()

# Ask user to select the first Excel file
file_path_1 = filedialog.askopenfilename(title="Select the 標準器校正‐実施計画_2022-2023 file")

# Load the first workbook
wb1 = openpyxl.load_workbook(file_path_1)# Print out the existing worksheets

# Print the list of worksheet names
wb1_names = wb1.sheetnames
print("Existing worksheets:")
for i, name in enumerate(wb1_names):
    print(f"{i+1}. {name}")

# Prompt the user to select a worksheet
ws1_num = int(input("Enter the number of the worksheet (SAP data): "))
# Get the selected worksheet
ws1 = wb1[wb1_names[ws1_num-1]]

# Prompt the user to select a worksheet
ws3_num = int(input("Enter the number of the worksheet (Cal. Date): "))
# Get the selected worksheet
ws3 = wb1[wb1_names[ws3_num-1]]

print("-"*47)

'''-------------------------------------------------------------------------------------------------------------------'''

# Ask user to select the second Excel file
file_path_2 = filedialog.askopenfilename(title="Select the SAP data file")

# Load the second workbook
wb2= openpyxl.load_workbook(file_path_2)# Print out the existing worksheets

# Print the list of worksheet names
wb2_names = wb2.sheetnames
print("Existing worksheets:")
for i, name in enumerate(wb2_names):
    print(f"{i+1}. {name}")

# Prompt the user to select a worksheet
ws2_num = int(input("Enter the number of the worksheet you want to use: "))
# Get the selected worksheet
ws2 = wb2[wb2_names[ws2_num-1]]    

print("-"*47)

'''-------------------------------------------------------------------------------------------------------------------'''
#DATA INSERTING

# Delete all data in column A
for cell in ws1['A']:
    cell.value = None

# Copy data from column A in sheet2 to sheet1
for index, cell in enumerate(ws2['C'], start=1):
    ws1.cell(row=index, column=1).value = cell.value

print("-"*47)

'''-------------------------------------------------------------------------------------------------------------------'''
#INSERTING ROW

# inserting the row for added number
for row in range(1, ws1.max_row + 1):
    if ws1.cell(row, column=1).value != ws1.cell(row, column=3).value:

        ws1.insert_rows(row)

        for row in range(row, ws1.max_row + 1):
            cell = ws1.cell(row+1, column=1)
            new_data_row = ws1.cell(row=row, column=1,value = cell.value)
            ws1.cell(row=row+1, column=1).value = None

print("-"*47)
'''-------------------------------------------------------------------------------------------------------------------'''
# INSERTING DATA

# get the last row and save as last_row
last_row = 0
for i, cell in enumerate(ws1['A'], start=1):
    if isinstance(cell.value, (int, float)):
        last_row = i

# append the number of added rows 
added_rows = []
for row in ws1.iter_rows(min_row=1, min_col=3, max_col=3,max_row=last_row):
    if row[0].value is None:
        added_rows.append(row)

# Loop through each row in column A of RAW_updated.xlsx
for i in range(1, ws1.max_row + 1):
    raw_val = ws1.cell(row=i, column=1).value  # Get the value in column A of RAW_updated.xlsx

    if ws1.cell(row=i, column=3).value is None:
        print(f"Column C in Row {i} in RAW_updated.xlsx is empty")

        # Search for the value in column C of SAP.xlsx
        for j in range(1, ws2.max_row + 1):
            step2_val = ws2.cell(row=j, column=3).value  # Get the value in column C of SAP.xlsx

            # If the values match, copy the data from column C to column Z
            if raw_val == step2_val:

                for k in range(3, 27):
                    ws1.cell(row=i, column=k).value = ws2.cell(row=j, column=k).value

print (f"Number of added rows in column C: {len(added_rows)}")

print("-"*47)

'''-------------------------------------------------------------------------------------------------------------------'''

counter = 1
for row in ws1.iter_rows(min_row=1, max_col=4, max_row=ws1.max_row):
    print(f"Row {counter}: Column A = {row[0].value}, Column C = {row[2].value}, Column D = {row[3].value}")
    counter += 1
print("-"*47)

'''-------------------------------------------------------------------------------------------------------------------'''
# INSERTING DATE DATA

# Loop through each row in column K
for cell in ws1['K']:
    date_string = cell.value
    if date_string is not None:
        # Split the date into day, month, and year
        day, month, year = date_string.split('.')

        # Write the parts to their respective columns in the same row
        row_number = cell.row
        ws3.cell(row=row_number, column=1, value=int(day))
        ws3.cell(row=row_number, column=2, value=int(month))
        ws3.cell(row=row_number, column=3, value=int(year))

counter = 1
for row in ws3.iter_rows(min_row=1, max_col=3, max_row=ws1.max_row):
    print(f"Row {counter}: Column A = {row[0].value}, Column B = {row[1].value}, Column C = {row[2].value}")
    counter += 1
print("-"*47)

'''-------------------------------------------------------------------------------------------------------------------'''
# Do some processing here... #time
end_time = time.time()
# Calculate the time elapsed
elapsed_time = end_time - start_time
# Convert elapsed time to minutes and seconds
minutes = int(elapsed_time // 60)
seconds = int(elapsed_time % 60)
print(f'The code took {minutes} minutes and {seconds} seconds to execute.')

def save_workbook(workbook, filename):
    workbook.save(filename)
save_workbook(wb1, "Intermediate_files.xlsx")